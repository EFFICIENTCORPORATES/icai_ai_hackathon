import pandas as pd


# Function to extract the first 2 digits of GSTIN and map to state name
def get_state_name(gstin):
    first_two_digits = gstin[:2]
    state_mapping = STATE_MAPPING
    return state_mapping.get(first_two_digits, 'Unknown')

def flatten_dict(dic):
    """
    This is  avery useful function to flatten any dictionary which consists of a nested list of a nested dictionary

    This function taken only 1 parameter ,

    The parameter must be a dictionary

    This function has a depndent function i.e expand_list

    Both these functions always to be used simultaneously as they are compimentary to each other


    """
        
    import pandas as pd
    import json
    import warnings
    from openpyxl import load_workbook
    import os

    warnings.filterwarnings('ignore')

    df2 = pd.DataFrame()

    key_list = list(dic.keys())
    flat_dict = dict()

    for i in key_list:
        dict_whole = {i: dic[i]}
        dict_value = dic[i]

        if isinstance(dict_value, dict):
            flat_dict.update(dict_value)

        elif isinstance(dict_value, list):

            if len(dict_value) == 1:
                a = dict_value[0]
                b = flatten_dict(a)
                flat_dict.update(b)
            elif len(dict_value) == 0:
                pass

            else:
                dicdf = expand_list(dict_value)
                flat_dict.update(dicdf)

        else:
            flat_dict.update(dict_whole)

    data_list = list(flat_dict.items())

    df = pd.DataFrame(data_list)
    df1 = df.T
    df1.columns = df1.loc[0]
    df1 = df1.drop(0)
    df1 = df1.reset_index(drop=True)
    return (flat_dict)

def expand_list(list_dic):
    """
    Flattens a list containing nested lists or dictionaries.

    Args:
        list_dic (list): A list potentially containing nested lists or dictionaries.

    Returns:
        dict: A flattened dictionary representation of the input list.
    """

    df_list = []  # Collect DataFrames for concatenation

    for i in list_dic:
        if isinstance(i, dict):
            flat_dictl = flatten_dict(i)  # Assuming `flatten_dict` is defined elsewhere
            try:
                df = pd.DataFrame(flat_dictl)
            except:
                df = pd.DataFrame(list(flat_dictl.items())).T
                df.columns = df.loc[0]
                df = df.drop(0).reset_index(drop=True)
            df_list.append(df)

        elif isinstance(i, list):
            df = expand_list(i)  # Recursively handle nested lists
            df_list.append(df)

        else:
            df = pd.DataFrame.from_dict({i: list_dic[i]}, orient="index").T
            df_list.append(df)

    df = pd.concat(df_list, ignore_index=True)  # Concatenate efficiently
    conv_dict = df.to_dict(orient="list")
    return conv_dict

def rename_r1_columns(dataframe):

    """

    This is a support function for the gstr2B to Excel conversion


    """
    
    dataframe.rename(columns={"idt":"Final_Invoice_CNDN_Date",
                "val":"Invoice_CNDN_Value",
                "rchrg":"Supply_Attract_Reverse_Charge",
                "itcavl":"ITC_Available",
                "diffprcnt":"Applicable_Percent_TaxRate",
                "pos":"Place_Of_Supply",
                "inv_typ":"Final_Inv_CNDN_Type",
                "inum":"Final_Invoice_CNDN_No",
                "rsn":"Reason",
                "samt":"SGST_Amount",
                "rt":"Tax_Rate",
                "num":"Check_num",
                "txval":"Taxable_Value",
                "camt":"CGST_Amount",
                "csamt":"Cess_Amount",
                "ctin":"GSTIN_of_Customer",
                "iamt":"IGST_Amount",
                "irn":"IRN",
                "cfs":"Counter_Party_Filing_Status",
                "cflag":"Counter_Party_Action",
                "updby":"Uploaded_By",
                "chksum":"Check_Sum",
                "flag":"Tax_Payer_Action",
                "irngendate":"IRN_Generate_Date",
                "srctyp":"Source_Type",
                "GSTR2B-Table":"GSTR2B-Table",
                "rtnprd":"GSTR2B_Period",
                "gstin":"Recipient_GSTIN",
                "Json File Name":"JSON_Source_File",
                "typ":"Final_Inv_CNDN_Type",
                "sply_ty":"Supply_Type",
                "nt_dt":"Final_Invoice_CNDN_Date",
                "nt_num":"Final_Invoice_CNDN_No",
                "d_flag":"d_flag",
                "exp_typ":"Final_Inv_CNDN_Type",
                "File_Name":"Source_Excel_File"},inplace=True)

def gstr1_to_excel(filepath):
    """
    This is a very easy to use funcion to extract the json data of GSTR1 into an excel file.

    This function takes only one argument i.e a completepath to the json file upto extension

    Simply pass the complete path and run.

    Invoice wise data will be populated in the Excel sheet

    """

    
    import pandas as pd
    import json
    import warnings
    from openpyxl import load_workbook
    import os

    warnings.filterwarnings('ignore')


    print(f'The Json GSTR-1 file path selected is {filepath}')
    print("We are analyzing the sheets available")


    folder = os.path.dirname(filepath)
    filename = "Converted_GSTR-1 Table Wise" + os.path.splitext(os.path.basename(filepath))[0] + ".xlsx"
    fullpath1 = os.path.join(folder, filename)
    
    pth = os.path.dirname(str(filepath))
    
    writer=pd.ExcelWriter(fullpath1, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_numbers': True}})

    os.makedirs(folder, exist_ok=True)
    
    # Constructing the filename using f-string
    filename = f"Summary_{os.path.splitext(os.path.basename(filepath))[0]}.xlsx"
    
    fullpath1a = os.path.join(folder, filename)

    writer1=pd.ExcelWriter(fullpath1a, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_numbers': True}})

    df1 = pd.DataFrame()
    df1.to_excel(writer1, sheet_name="Summary_GSTR1", index=False)

    writer1.close()

    fullpath2 = fullpath1a  # this is a very useful command for defining the correct filepath
    print("fulpath2: ",fullpath2)
    wb = load_workbook(fullpath2)
    ws = wb["Summary_GSTR1"]

    ws["A1"].value = "AUTOMATION SOLUTIONS BY EFFICIENT CORPORATES-[TM]"
    ws["A4"].value = "Summary of the GSTR-1 File Compiled"

    ws["A6"].value = "GSTIN of the Seller"
    ws["A7"].value = "GSTR-1 filing period"
    ws["A8"].value = "GSTR-1 filing Type"
    ws["A9"].value = "GT"

    ws["A10"].value = "CUR_GT"
    ws["A11"].value = "FILING DATE"

    ws["A15"].value = "SUMMARY OF THE DIFFERENT TABLES IN THE GSTR-1 FILE"

    ws.merge_cells("A15:I15")

    ws["A17"].value = "GSTR-1 Tables"
    ws["B17"].value = "Count"
    ws["C17"].value = "Invoice Amount"
    ws["D17"].value = "Taxable Value"
    ws["E17"].value = "IGST"
    ws["F17"].value = "CGST"

    ws["G17"].value = "SGST"
    ws["H17"].value = "Cess"
    ws["I17"].value = "Total Tax Amount"

    ws["A18"].value = "BUSINESS-2- BUSINESS (B2B)"
    ws["A19"].value = "BUSINESS-2- CONSUMER-SMALL (B2CS)"

    ws["A20"].value = "BUSINESS-2- CONSUMER-LARGE (B2CL)"
    ws["A21"].value = "EXPORT (EXP)"
    ws["A22"].value = "CREDIT NOTE / DEBIT NOTE (CDNR)"

    ws["A24"].value = "HSN SUMMARY"
    
    with open(filepath) as json_file:
        data = json.load(json_file)

    dic_keys = data.keys
    
    df_all_combined=pd.DataFrame()

    for i in dic_keys():

        if i == "gstin":
            print("We are getting the Meta Data for you...Please wait...!")
            gst = data[i]
            ws["B6"].value = gst

        elif i == "fp":
            fp = data[i]
            ws["B7"].value = fp

        elif i == "filing_typ":
            fil = data[i]
            if fil == "M":
                ws["B8"].value = "Monthly"
            elif fil == "Q":
                ws["B8"].value = "Quarterly"
            else:
                ws["B8"].value = fil

        elif i == "gt":
            gt = data[i]
            ws["B9"].value = gt

        elif i == "cur_gt":
            cur_gt = data[i]
            ws["B10"].value = cur_gt

        elif i == "b2b":

            print("Fetching the B2B data, Please wait for some time...!!")
            b2b_data = data[i]
            dic_b2b = expand_list(b2b_data)
            
            try:
                df_b2b = pd.DataFrame(dic_b2b)
            except ValueError:
                df_b2b = pd.DataFrame(dic_b2b, index=[0])

            df_b2b["GSTR1-Table"] = "B2B"
            # df_b2b["Json File Name"] = filepath
            df_b2b.to_excel(writer, sheet_name='B2B_DATA', index=False)
            df_all_combined = pd.concat([df_all_combined, df_b2b], ignore_index=True)

        elif i == "b2cl":

            print("Fetching the B2CL data, Please wait for some time...!!")
            b2cl_data = data[i]
            dic_b2cl = expand_list(b2cl_data)


            try:
                df_b2cl = pd.DataFrame(dic_b2cl)
            except ValueError:
                df_b2cl = pd.DataFrame(dic_b2cl, index=[0])

            # df_b2cl = pd.DataFrame(dic_b2cl)
            df_b2cl["GSTR1-Table"] = "B2C-L"
            # df_b2cl["Json File Name"] = filepath
            
            df_b2cl.to_excel(writer, sheet_name='B2CL_DATA', index=False)
            
            df_all_combined = pd.concat([df_all_combined, df_b2cl], ignore_index=True)

        elif i == "cdnr":

            print("Fetching the CDNR data, Please wait for some time...!!")
            cdnr_data = data[i]
            dic_cdnr = expand_list(cdnr_data)

            try:
                df_cdnr = pd.DataFrame(dic_cdnr)
            except ValueError:
                df_cdnr = pd.DataFrame(dic_cdnr, index=[0])

            df_cdnr["GSTR1-Table"] = "CDNR"
            # df_cdnr["Json File Name"] = filepath
            
            df_cdnr.to_excel(writer, sheet_name='CDNR_DATA', index=False)
            
            df_cdnr.rename(columns={"inv_typ":"Note_Supply_Type",
                "ntty":"Final_Inv_CNDN_Type"},inplace=True)
            
            df_all_combined = pd.concat([df_all_combined, df_cdnr], ignore_index=True)


        elif i == "exp":

            print("Fetching the Export data, Please wait for some time...!!")
            exp_data = data[i]
            dic_exp = expand_list(exp_data)

            try:
                df_exp = pd.DataFrame(dic_exp)
            except ValueError:
                df_exp = pd.DataFrame(dic_exp, index=[0])


            # df_exp = pd.DataFrame(dic_exp)
            df_exp["GSTR1-Table"] = "EXPORT"
            # df_exp["Json File Name"] = filepath
            
            df_exp.to_excel(writer, sheet_name='EXPORT_DATA', index=False)
            
            df_all_combined = pd.concat([df_all_combined, df_exp], ignore_index=True)

        elif i == "b2cs":

            print("Fetching the B2CS data, Please wait for some time...!!")
            b2cs_data = data[i]
            dic_b2cs = expand_list(b2cs_data)

            try:
                df_b2cs = pd.DataFrame(dic_b2cs)
            except ValueError:
                df_b2cs = pd.DataFrame(dic_b2cs, index=[0])
            df_b2cs["GSTR1-Table"] = "B2C-S"
            # df_b2cs["Json File Name"] = filepath
            
            df_b2cs.to_excel(writer, sheet_name='B2CS_DATA', index=False)
            
            
            df_all_combined = pd.concat([df_all_combined, df_b2cs], ignore_index=True)

        elif i == "hsn":

            print("Getting the HSN Summary For you...!!")
            hsn_data = data[i]
            dic_hsn = flatten_dict(hsn_data)

            try:
                df_hsn = pd.DataFrame(dic_hsn)
            except ValueError:
                df_hsn = pd.DataFrame(dic_hsn, index=[0])

            df_hsn.to_excel(writer, sheet_name='HSN_DATA', index=False)

        elif i == "nil":
            nil_data = data[i]
            dic_nil = flatten_dict(nil_data)

            try:
                df_nil = pd.DataFrame(dic_nil)
            except ValueError:
                df_nil = pd.DataFrame(dic_nil, index=[0])

            df_nil.to_excel(writer, sheet_name='NIL_NONGST_DATA', index=False)

        elif i == "doc_issue":
            print("Getting the Document Series Summary For you...!!")
            doc_data = data[i]
            dic_doc = flatten_dict(doc_data)

            try:
                df_doc = pd.DataFrame(dic_doc)
            except ValueError:
                df_doc = pd.DataFrame(dic_doc, index=[0])
            df_doc.to_excel(writer, sheet_name='DOC_SERIES_DATA', index=False)

        elif i == "fil_dt":
            fildt = data["fil_dt"]
            ws["B11"].value = fildt

        else:
            add_case = data[i]
            
            if isinstance(add_case, list):
                dic_add_case = expand_list(add_case)

                try:
                    df_add_case = pd.DataFrame(dic_add_case)
                except ValueError:
                    df_add_case = pd.DataFrame(dic_add_case, index=[0])


                # df_add_case = pd.DataFrame(dic_add_case)
                df_add_case["GSTR1-Table"] = i
                df_add_case.to_excel(writer, sheet_name=i, index=False)
            
            elif isinstance(add_case, dict):
                dic_add_case = flatten_dict(add_case)

                try:
                    df_add_case = pd.DataFrame(dic_add_case)
                except ValueError:
                    df_add_case = pd.DataFrame(dic_add_case, index=[0])
                df_add_case["GSTR1-Table"] = i
                df_add_case.to_excel(writer, sheet_name=i, index=False)
            else:
                pass
    print(data.keys())
    
    print("Consolidating All Major Tables in Single Sheet for you..!!")
    
    rename_r1_columns(df_all_combined)
    df_all_combined.to_excel(writer, sheet_name="effcorp_all_combined", index=False)

    wb.save(fullpath2)
    writer.close()

    try:
        ws["B18"].value = len(df_b2b["ctin"])
        ws["C18"].value = df_b2b["val"].sum()
        ws["D18"].value = df_b2b["txval"].sum()
        ws["E18"].value = df_b2b["iamt"].sum()
        ws["F18"].value = df_b2b["camt"].sum()
        ws["G18"].value = df_b2b["samt"].sum()
        ws["H18"].value = df_b2b["csamt"].sum()
        ws["I18"].value = df_b2b["iamt"].sum() + df_b2b["camt"].sum() + df_b2b["samt"].sum() + df_b2b["csamt"].sum()
    except:
        pass

    try:

        ws["B19"].value = len(df_b2cs["rt"])
        #     ws["C19"].value = sum(df_b2cs["val"])
        ws["D19"].value = df_b2cs["txval"].sum()
        ws["E19"].value = df_b2cs["iamt"].sum()
        ws["F19"].value = df_b2cs["camt"].sum()
        ws["G19"].value = df_b2cs["samt"].sum()
        # ws["H19"].value = df_b2cs["csamt"].sum()
        ws["I19"].value = df_b2cs["iamt"].sum() + df_b2cs["camt"].sum() + df_b2cs["samt"].sum()
    except:
        pass

    try:
        ws["B20"].value = len(df_b2cl["val"])
        ws["C20"].value = df_b2cl["val"].sum()
        ws["D20"].value = df_b2cl["txval"].sum()
        ws["E20"].value = df_b2cl["iamt"].sum()
        #     ws["F20"].value = df_b2cl["camt"].sum()
        #     ws["G20"].value = df_b2cl["samt"].sum()
        ws["H20"].value = df_b2cl["csamt"].sum()
        ws["I20"].value = df_b2cl["iamt"].sum()

    except:
        pass


    try:

        ws["B21"].value = len(df_exp["flag"])
        ws["C21"].value = df_exp["val"].sum()
        ws["D21"].value = df_exp["txval"].sum()
        ws["E21"].value = df_exp["iamt"].sum()
        #     ws["F21"].value = df_exp["camt"].sum()
        #     ws["G21"].value = df_exp["samt"].sum()
        ws["H21"].value = df_exp["csamt"].sum()
        ws["I21"].value = df_exp["iamt"].sum()
    except:
        pass


    try:

        ws["B22"].value = len(df_cdnr["flag"])
        ws["C22"].value = df_cdnr["val"].sum()
        ws["D22"].value = df_cdnr["txval"].sum()
        ws["E22"].value = df_cdnr["iamt"].sum()
        #     ws["F22"].value = df_cdnr["camt"].sum()
        #     ws["G22"].value = df_cdnr["samt"].sum()
        ws["H22"].value = df_cdnr["csamt"].sum()
        ws["I22"].value = df_cdnr["iamt"].sum()

    except:
        pass

    try:
        ws["B24"].value = len(df_hsn["flag"])
        #     ws["C24"].value = df_hsn["val"].sum()
        ws["D24"].value = df_hsn["txval"].sum()
        ws["E24"].value = df_hsn["iamt"].sum()
        ws["F24"].value = df_hsn["camt"].sum()
        ws["G24"].value = df_hsn["samt"].sum()
        ws["H24"].value = df_hsn["csamt"].sum()
        ws["I24"].value = df_hsn["iamt"].sum() + df_hsn["camt"].sum() + df_hsn["samt"].sum() + df_hsn["csamt"].sum()

    except:
        pass

    writer.close()

    print("All Data have been extracted Successfully! ")

    wb.save(fullpath2)
    writer.close()

    wb.close()
    writer.close()

    print("We have created two Excel files for you..!! 1) Summary.xlsx and 2) GSTR-1 Table Wise.xlsx")
    # print(f'The Excel Files are Extracted and kept in the below path \n {fullpath2}\n{fullpath1}\n\n ')
    sg.popup(f'The Excel Files are Extracted and kept in the below path \n {fullpath2}\n\n ')
    return(fullpath1)

def rename_columns(df):
    # Define column name mappings for B2B and CDN sections

    df.rename(columns={
        "ctin": "Counterparty_GSTIN",
        "cfs": "CFS",
        "cfs3b": "CFS3B",
        "val": "Invoice_Value",
        "inv_typ": "Invoice_Type",
        "pos": "Place_of_Supply",
        "srctyp": "Source_Type",
        "idt": "Invoice_Date",
        "rchrg": "Reverse_Charge",
        "inum": "Invoice_Number",
        "chksum": "Checksum",
        "cfs": "CFS",
        "ctin": "Counterparty_GSTIN",
        "nt": "Notes",
        "val": "Note_Value",
        "irn": "IRN",
        "d_flag": "Debit_Note_Flag",
        "nt_num": "Note_Number",
        "irngendate": "Note_Generation_Date",
        "nt_dt": "Note_Date",
        "inv_typ": "Invoice_Type",
        "pos": "Place_of_Supply",
        "srctyp": "Source_Type",
        "ntty": "Note_Type",
        "rchrg": "Reverse_Charge",
        "chksum": "Checksum"
                },inplace=True)

